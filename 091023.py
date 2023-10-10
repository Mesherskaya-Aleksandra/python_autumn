# import collections
# s=input()
# d=collections.Counter(s)
# for k,v in d.items():
#     print(k,v)
#
# import collections
# s=['a', 'b', 1,2,3]
# d=collections.Counter(s)
# for k,v in d.items():
#     print(k,v)
#
# import collections
# s='jglihufhidlasfh'
# d=collections.Counter(s)
# print(dict(d))

# dct={1:{11:{111:1111}}, 2:{22:{222:2222}},3:{33:{333}, 4:444}}
# for k, v in dct.items():
#     print(k,v)
#     if type(v)==dict:
#         for a,b in v.items():
#             print('    ',a,b)
#             if type(b)==dict:
#                 for x, y in b.items():
#                     print('        ',x,y)

#
# f=open('mmm.txt', encoding='utf-8')
# s=f.read()
# for i in s[:20]:
#     print(i,end='')
# f.close()
#
# f=open('test.txt', encoding='utf-8')
# s=f.read()
# for i in s[:20]:
#     print(i,end='')
# f.close()

# f=open('test.txt', encoding='utf-8')
# s=f.read()
# for i in s[:20]:
#     print(i,end='')
# f.close()
#
# f=open("test.txt", encoding='utf-8' )
# s=f.read(10)
# p=f.read(10)
# print(s,p,sep='----------')
# f.close()

# f=open("test.txt", encoding='utf-8' )
# for i in range(5):
#     s=f.read(10)
#     print(s)
# f.close()


# f=open("test.txt", encoding='utf-8' )
# for i in range(5):
#     s=f.read(10)
#     print(s.strip())
# f.close()


# f=open("test.txt", encoding='utf-8' )
# s=f.read()
# print(s)
# f.close()

# f=open("test.txt", encoding='utf-8' )
# s=f.readlines()
# print(s)
# f.close()
#
# f=open("test.txt", encoding='utf-8' )
# s=f.readlines()
# print(type(s))
# f.close()

# f=open("test.txt", encoding='utf-8' )
# s=f.readlines()
# for i in s:
#     print(i)
# f.close()

# f=open("test.txt", encoding='utf-8' )
# s=f.readlines()
# for i in s:
#     print(i.strip())
# f.close()

# f=open("test.txt", encoding='utf-8' )
# s=f.readlines()
# print(s[-1])
# f.close()

# f=open("test.txt", encoding='utf-8' )
# lst=f.readlines()
# for i,j in enumerate(lst,1):
#     print(i,j.strip('\n'))
# f.close()

# f=open("test.txt", encoding='utf-8' )
# s=f.readline(2)
# print(s.strip( ))
# f.close()

# f=open("test.txt", encoding='utf-8' )
# lst=list(f.readline())
# for i, j in enumerate(lst):
#     print(i,j)
# f.close()

# f=open("test.txt", encoding='utf-8' )
# i=1
# for i in range(1,5):
#     s=f.readline()
#     print(i,s.strip())
# f.close()

# f=open("test.txt", encoding='utf-8' )
# i=1
# for i in range(5,9):
#     s=f.readline()
#     print(i,s.strip())
# f.close()


# f=open("test.txt", mode = 'w', encoding='utf-8' )
# f.write('The first line\n')
# f.write('fgsar'+'\n')
# f.write('Cesar\n')
# a=['Первый элемент списка\n', 'Второй элемент списка']
# f.writelines(a)
# f.close()

# g=open("test.txt", mode = 'r', encoding='utf-8' )
# for i in g:
#     print(i.strip())
# g.close()

# f=open('test.txt', mode='r', encoding='utf-8')
# s=f.read()
# print(s)
# f.close()
#
# g=open('test.txt', mode='w', encoding='utf-8')
# g.write(s)
# g.close()

# s=f.readilnes()
# print(s)
# f.close

# g=open('test.txt', mode='w', encoding='utf-8')
# for i in s:
#     for j in i:
#         if j.isdigit():
#             g.write(i)
#             break
# g.close()

# import openpyxl
# from openpyxl import Workbook
# wb=Workbook()
# wb.save('test.xlsx')

# import openpyxl
# wb=openpyxl.load_workbook('test.xlsx')
# print(wb.sheetnames)
# ws=wb.active
# print(ws.title)
# wb.create_sheet('List1')
# ws3=wb['List1']
# print(ws3.title)
#
# print(wb.sheetnames)
# wb.active=wb['List1']
# wb.remove(ws)
# print(wb.sheetnames)
#
# ws3['A1'].value=345
# s=ws3['A1'].value
# print(s)
#
# ws3['A1'].value=1000000000000
# s=ws3['A1'].value
# print(s)
#
# ws3['B2'].value='dfasfas'
# wb.save('test.xlsx')
#
# print(ws3.cell(row=1, column=2).value)

import openpyxl
from openpyxl import Workbook
wb=Workbook()
wb.save('test1.xlsx')

import openpyxl
wb=openpyxl.load_workbook('test1.xlsx')
ws=wb.active
i,j=1,1
for i in range(1,2):
    for j in range(1,3):
        print(i, j, ws.cell(row=i,column=1).value)
# for i in range(5,8):
#     ws.cell(row=i)

for i in range(1, ws.max_row):
    print(i, ws.cell(row=i, column=1).value)